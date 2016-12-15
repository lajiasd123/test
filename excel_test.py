# coding:UTF-8
from  openpyxl.reader.excel  import  load_workbook
import openpyxl, pprint,os


#TODO:打开一份excel文件，获取sheet1页的内容
wjnr = openpyxl.load_workbook('1.xlsx')
sheet = wjnr.get_sheet_names()
print '该文件包含以下几张表，请输入所要操作的表(1-'+str(len(sheet))+')'
print sheet
sheetinput = raw_input('>>>')
sheetname = 'Sheet' + sheetinput
print '正在打开'+ sheetname +'...'
bnr = wjnr.get_sheet_by_name(sheetname)
total = 0
for row in range(1,bnr.max_row + 1):
	liea = bnr['A'+str(row)].value
	lieb = bnr['B'+str(row)].value
	total += int(lieb)

#TODO：将结果保存进一个txt文件并进行存储
print '开始写入文件...'
excel_testfile = open('excel.txt','w')
excel_testfile.write('和是' + str(total))
excel_testfile.close()
print '写入完毕'
